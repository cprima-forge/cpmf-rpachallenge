"""Generic Excel file source - schema-driven parsing.

Generic - no domain-specific imports.
"""

from pathlib import Path
from typing import Any, Callable, Iterable, Mapping

import openpyxl

from ..schema import Schema

RowToDict = Callable[[tuple, Schema], Mapping[str, Any]]


def default_row_to_dict(row: tuple, schema: Schema, header_map: dict[str, int] | None = None) -> dict[str, Any]:
    """Default row parser using schema.

    Args:
        row: Row values tuple
        schema: Schema defining columns
        header_map: Optional mapping of column name to index (for header-based mapping)

    Returns:
        Dictionary mapping column names to parsed values
    """
    if header_map:
        # Map by column name using header positions
        return {
            col.name: col.parse(row[header_map[col.name]])
            for col in schema
            if col.name in header_map and header_map[col.name] < len(row)
        }
    else:
        # Legacy: map by position (assumes schema order matches Excel order)
        return {
            col.name: col.parse(row[i]) for i, col in enumerate(schema) if i < len(row)
        }


class XlsxSource:
    """Generic Excel file source - schema-driven parsing."""

    def __init__(
        self,
        path: Path | str,
        schema: Schema,
        row_to_dict: RowToDict = default_row_to_dict,
        *,
        data_start_row: int = 2,
        use_headers: bool = True,
    ):
        self.path = Path(path)
        self.schema = schema
        self.row_to_dict = row_to_dict
        self.data_start_row = data_start_row
        self.use_headers = use_headers
        self._header_map: dict[str, int] | None = None

    def _build_header_map(self, header_row: tuple) -> dict[str, int]:
        """Build mapping of column names to indices from header row.

        Maps Excel column names to schema field names:
        - "First Name" -> "first_name"
        - "Phone Number" -> "phone"
        etc.
        """
        # Normalize Excel headers to match schema field names
        excel_to_schema = {
            "First Name": "first_name",
            "Last Name": "last_name",
            "Phone Number": "phone",
            "Email": "email",
            "Address": "address",
            "Company Name": "company_name",
            "Role in Company": "role",
        }

        header_map = {}
        for idx, header in enumerate(header_row):
            # Trim whitespace from header
            clean_header = header.strip() if isinstance(header, str) else header
            if clean_header in excel_to_schema:
                schema_name = excel_to_schema[clean_header]
                header_map[schema_name] = idx

        return header_map

    def load(self) -> Iterable[dict[str, Any]]:
        """Yield records from Excel file."""
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.active

        # Read header row if using header-based mapping
        if self.use_headers:
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            self._header_map = self._build_header_map(header_row)

        for row in sheet.iter_rows(min_row=self.data_start_row, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            yield self.row_to_dict(row, self.schema, self._header_map)
