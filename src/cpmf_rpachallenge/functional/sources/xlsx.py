"""Generic Excel file source - schema-driven parsing.

Generic - no domain-specific imports.
"""

from pathlib import Path
from typing import Any, Callable, Iterable, Mapping

import openpyxl

from ..schema import Schema

RowToDict = Callable[[tuple, Schema], Mapping[str, Any]]


def default_row_to_dict(row: tuple, schema: Schema, header_map: dict[str, int] | None = None) -> dict[str, Any]:
    """Default row parser using schema.

    Args:
        row: Row values tuple
        schema: Schema defining columns
        header_map: Optional mapping of column name to index (for header-based mapping)

    Returns:
        Dictionary mapping column names to parsed values
    """
    if header_map:
        # Map by column name using header positions
        return {
            col.name: col.parse(row[header_map[col.name]])
            for col in schema
            if col.name in header_map and header_map[col.name] < len(row)
        }
    else:
        # Legacy: map by position (assumes schema order matches Excel order)
        return {
            col.name: col.parse(row[i]) for i, col in enumerate(schema) if i < len(row)
        }


class XlsxSource:
    """Generic Excel file source - schema-driven parsing.

    Domain-agnostic - requires header_map to be passed in for header-based mapping.
    """

    def __init__(
        self,
        path: Path | str,
        schema: Schema,
        row_to_dict: RowToDict = default_row_to_dict,
        *,
        data_start_row: int = 2,
        use_headers: bool = True,
        header_map: dict[str, str] | None = None,
    ):
        """Initialize Excel source.

        Args:
            path: Path to Excel file
            schema: Schema defining columns
            row_to_dict: Function to convert row tuple to dict
            data_start_row: Row number where data starts (1-indexed)
            use_headers: Whether to use header-based mapping
            header_map: Optional mapping of Excel headers to schema field names
                       (e.g., {"First Name": "first_name"})
        """
        self.path = Path(path)
        self.schema = schema
        self.row_to_dict = row_to_dict
        self.data_start_row = data_start_row
        self.use_headers = use_headers
        self.excel_header_map = header_map
        self._column_indices: dict[str, int] | None = None

    def _build_column_indices(self, header_row: tuple) -> dict[str, int]:
        """Build mapping of schema field names to column indices.

        Uses excel_header_map to convert Excel headers to schema field names.

        Args:
            header_row: Tuple of Excel header values

        Returns:
            Dict mapping schema field name to column index
        """
        column_indices = {}

        if not self.excel_header_map:
            # No header map provided - use direct schema field name matching
            for idx, header in enumerate(header_row):
                clean_header = header.strip() if isinstance(header, str) else header
                # Try to find matching schema field
                for col in self.schema:
                    if col.name == clean_header or col.name.replace("_", " ").lower() == str(clean_header).lower():
                        column_indices[col.name] = idx
                        break
        else:
            # Use provided header map
            for idx, header in enumerate(header_row):
                clean_header = header.strip() if isinstance(header, str) else header
                if clean_header in self.excel_header_map:
                    schema_name = self.excel_header_map[clean_header]
                    column_indices[schema_name] = idx

        return column_indices

    def load(self) -> Iterable[dict[str, Any]]:
        """Yield records from Excel file."""
        wb = openpyxl.load_workbook(self.path)
        sheet = wb.active

        # Read header row if using header-based mapping
        if self.use_headers:
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            self._column_indices = self._build_column_indices(header_row)

        for row in sheet.iter_rows(min_row=self.data_start_row, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            yield self.row_to_dict(row, self.schema, self._column_indices)
